import smtplib
import os
import html as html_lib
import requests
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import database
import analytics

def format_duration(seconds):
    """Преобразует секунды в читаемый вид (ЧЧ:ММ:СС)."""
    if seconds is None:
        return "-"
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"

def build_report_summary(start_time, end_time):
    """
    Собирает данные отчета в структуре печатной версии (PDF):
    сводные KPI, топ-20 проблемных серверов и список всех проблем по серверам.
    """
    report_data = analytics.calculate_sli_report(start_time, end_time)
    comments = database.get_incident_comments()

    servers = []
    incidents = []
    mttd_sum, mttd_cnt = 0, 0
    mttr_sum, mttr_cnt = 0, 0
    total_incidents = 0

    for client_name, client_data in report_data.items():
        if client_data['mttd_avg_sec'] > 0:
            mttd_sum += client_data['mttd_avg_sec']
            mttd_cnt += 1
        if client_data['mttr_avg_sec'] > 0:
            mttr_sum += client_data['mttr_avg_sec']
            mttr_cnt += 1
        total_incidents += client_data['total_incidents_count']

        for srv in client_data['servers']:
            servers.append({
                'client': client_name,
                'name': srv['name'],
                'sla_percent': srv['sla_percent'],
                'downtime_sec': srv['downtime_sec'],
                'incidents_count': srv['incidents_count'],
                'comment': srv['comment'] or ''
            })
            for inc in srv['incidents']:
                if inc.get('is_ignored_by_pattern'):
                    excluded = 'исключено правилом'
                elif inc['is_maintenance']:
                    excluded = 'обслуживание'
                elif inc.get('is_power_issue'):
                    excluded = 'электропитание'
                elif inc['is_vpn_issue']:
                    excluded = 'сеть/VPN'
                else:
                    excluded = ''
                incidents.append({
                    'client': client_name,
                    'server': srv['name'],
                    'name': inc['name'],
                    'clock': inc['clock'],
                    'downtime_sec': inc.get('downtime_in_period_sec', 0) or 0,
                    'excluded': excluded,
                    'comment': comments.get(inc['eventid'], {}).get('comment', '')
                })

    # Средний SLA — по всем серверам парка (единая формула с дашбордом и ТВ-панелью)
    avg_sla = round(sum(s['sla_percent'] for s in servers) / len(servers), 3) if servers else 100.0
    top_servers = sorted(servers, key=lambda s: (s['sla_percent'], -s['downtime_sec']))[:20]
    incidents.sort(key=lambda i: (i['client'], i['server'], -i['clock']))

    return {
        'avg_sla': avg_sla,
        'total_incidents': total_incidents,
        'avg_mttd': int(mttd_sum / mttd_cnt) if mttd_cnt else 0,
        'avg_mttr': int(mttr_sum / mttr_cnt) if mttr_cnt else 0,
        'servers_total': len(servers),
        'top_servers': top_servers,
        'incidents': incidents
    }

def _format_incident_moment(clock):
    return datetime.datetime.fromtimestamp(clock).strftime('%d.%m.%Y %H:%M')

def generate_excel_report(start_time, end_time):
    """
    Генерирует Excel-отчет в структуре печатной версии (PDF):
    KPI, топ-20 проблемных серверов, список всех проблем по серверам.
    """
    summary = build_report_summary(start_time, end_time)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SLI-SLA Report"

    # Стили
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    section_font = Font(name='Segoe UI', size=12, bold=True, color='000000')
    normal_font = Font(name='Segoe UI', size=11)
    bold_font = Font(name='Segoe UI', size=11, bold=True)
    small_font = Font(name='Segoe UI', size=10, color='555555')

    blue_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    light_blue_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

    border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Шапка отчета
    ws.merge_cells('A1:G2')
    title_cell = ws['A1']
    title_cell.value = f"Отчет по доступности серверов (SLI/SLA)\nПериод: {datetime.datetime.fromtimestamp(start_time).strftime('%d.%m.%Y')} - {datetime.datetime.fromtimestamp(end_time).strftime('%d.%m.%Y')}"
    title_cell.font = title_font
    title_cell.fill = blue_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Сводные KPI
    kpi_pairs = [
        ("Средний SLA (по серверам)", f"{summary['avg_sla']}%"),
        ("Всего инцидентов", summary['total_incidents']),
        ("Ср. время реакции (MTTD)", format_duration(summary['avg_mttd'])),
        ("Ср. время решения (MTTR)", format_duration(summary['avg_mttr'])),
    ]
    row_num = 4
    for i, (label, value) in enumerate(kpi_pairs):
        label_cell = ws.cell(row=row_num + i, column=1, value=label)
        label_cell.font = bold_font
        label_cell.fill = light_blue_fill
        label_cell.border = thin_border
        value_cell = ws.cell(row=row_num + i, column=2, value=value)
        value_cell.font = normal_font
        value_cell.border = thin_border
        value_cell.alignment = Alignment(horizontal='right')

    # Раздел: Топ-20 проблемных серверов
    row_num = 9
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    sec_cell = ws.cell(row=row_num, column=1, value="Топ-20 проблемных серверов (по SLA)")
    sec_cell.font = section_font

    row_num += 1
    top_headers = ["№", "Клиент", "Сервер", "Uptime %", "Время простоя", "Сбоев", "Комментарий"]
    for col_num, header in enumerate(top_headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[row_num].height = 24

    for idx, srv in enumerate(summary['top_servers'], 1):
        row_num += 1
        values = [
            idx, srv['client'], srv['name'], f"{srv['sla_percent']}%",
            format_duration(srv['downtime_sec']), srv['incidents_count'], srv['comment']
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = bold_font if col_num == 2 else normal_font
            cell.border = thin_border
            if col_num in (1, 5, 6):
                cell.alignment = Alignment(horizontal='center')
            elif col_num == 4:
                cell.alignment = Alignment(horizontal='right')

    # Раздел: Все проблемы по серверам за период
    row_num += 2
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    sec_cell = ws.cell(row=row_num, column=1, value="Все проблемы по серверам за период")
    sec_cell.font = section_font

    row_num += 1
    inc_headers = ["Клиент", "Сервер", "Инцидент", "Начало", "Время простоя", "Категория", "Комментарий"]
    for col_num, header in enumerate(inc_headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[row_num].height = 24

    if not summary['incidents']:
        row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        ws.cell(row=row_num, column=1, value="За выбранный период проблем не зафиксировано").font = normal_font
    else:
        for inc in summary['incidents']:
            row_num += 1
            values = [
                inc['client'], inc['server'], inc['name'], _format_incident_moment(inc['clock']),
                format_duration(inc['downtime_sec']), inc['excluded'] or '—', inc['comment']
            ]
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = bold_font if col_num == 1 else (small_font if col_num in (4, 6) else normal_font)
                cell.border = thin_border
                if col_num in (4, 5, 6):
                    cell.alignment = Alignment(horizontal='center')

    # Ширина колонок
    widths = {'A': 26, 'B': 30, 'C': 45, 'D': 18, 'E': 16, 'F': 16, 'G': 40}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def send_email_report(recipient, subject, start_time, end_time):
    """
    Отправляет email с HTML-отчетом (в структуре печатной версии) и вложенным Excel-файлом.
    """
    settings = database.get_settings()

    # Проверка настроек SMTP
    smtp_host = settings.get('smtp_host')
    smtp_port = int(settings.get('smtp_port', '587'))
    smtp_user = settings.get('smtp_user')
    smtp_password = settings.get('smtp_password')
    smtp_from = settings.get('smtp_from', smtp_user)
    use_tls = settings.get('smtp_use_tls') == '1'

    if not smtp_host or not smtp_user:
        raise Exception("SMTP-сервер не настроен в системе.")

    summary = build_report_summary(start_time, end_time)
    esc = html_lib.escape

    date_start_str = datetime.datetime.fromtimestamp(start_time).strftime('%d.%m.%Y')
    date_end_str = datetime.datetime.fromtimestamp(end_time).strftime('%d.%m.%Y')

    sla_color = '#166534' if summary['avg_sla'] >= 99.5 else ('#9a3412' if summary['avg_sla'] >= 98.0 else '#991b1b')

    html_body = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #333; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 10px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px 10px; text-align: left; font-size: 13px; }}
            th {{ background-color: #1F4E78; color: white; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
            h3 {{ margin-top: 28px; margin-bottom: 4px; }}
            .kpi-table td {{ text-align: center; }}
            .kpi-label {{ font-size: 11px; color: #777; text-transform: uppercase; }}
            .kpi-value {{ font-size: 20px; font-weight: bold; }}
            .sla-good {{ color: green; font-weight: bold; }}
            .sla-warn {{ color: #b45309; font-weight: bold; }}
            .sla-bad {{ color: red; font-weight: bold; }}
            .inc-meta {{ font-size: 11px; color: #777; }}
            .footer {{ margin-top: 20px; font-size: 12px; color: #777; }}
        </style>
    </head>
    <body>
        <h2>Отчет по доступности ИТ-сервисов (SLA/SLI)</h2>
        <p>За период: <b>{date_start_str} — {date_end_str}</b></p>

        <table class="kpi-table">
            <tr>
                <td><div class="kpi-label">Средний SLA</div><div class="kpi-value" style="color: {sla_color};">{summary['avg_sla']}%</div></td>
                <td><div class="kpi-label">Всего инцидентов</div><div class="kpi-value">{summary['total_incidents']}</div></td>
                <td><div class="kpi-label">Ср. время реакции (MTTD)</div><div class="kpi-value">{format_duration(summary['avg_mttd'])}</div></td>
                <td><div class="kpi-label">Ср. время решения (MTTR)</div><div class="kpi-value">{format_duration(summary['avg_mttr'])}</div></td>
            </tr>
        </table>

        <h3>Топ-20 проблемных серверов (по SLA)</h3>
        <table>
            <thead>
                <tr>
                    <th style="width: 30px; text-align: center;">№</th>
                    <th>Клиент</th>
                    <th>Сервер</th>
                    <th style="text-align: right;">Uptime %</th>
                    <th style="text-align: center;">Время простоя</th>
                    <th style="text-align: center;">Сбоев</th>
                    <th>Комментарий</th>
                </tr>
            </thead>
            <tbody>
    """

    if not summary['top_servers']:
        html_body += '<tr><td colspan="7" style="text-align: center;">Нет данных</td></tr>'
    else:
        for idx, srv in enumerate(summary['top_servers'], 1):
            sla_class = 'sla-good' if srv['sla_percent'] >= 99.5 else ('sla-warn' if srv['sla_percent'] >= 98.0 else 'sla-bad')
            html_body += f"""
            <tr>
                <td style="text-align: center;">{idx}</td>
                <td><b>{esc(srv['client'])}</b></td>
                <td>{esc(srv['name'])}</td>
                <td style="text-align: right;" class="{sla_class}">{srv['sla_percent']}%</td>
                <td style="text-align: center;">{format_duration(srv['downtime_sec'])}</td>
                <td style="text-align: center;">{srv['incidents_count']}</td>
                <td class="inc-meta">{esc(srv['comment']) or '-'}</td>
            </tr>
            """

    html_body += """
            </tbody>
        </table>

        <h3>Все проблемы по серверам за период</h3>
        <table>
            <thead>
                <tr>
                    <th>Клиент</th>
                    <th>Сервер</th>
                    <th>Инцидент</th>
                    <th>Комментарий</th>
                    <th style="text-align: center;">Время простоя</th>
                </tr>
            </thead>
            <tbody>
    """

    if not summary['incidents']:
        html_body += '<tr><td colspan="5" style="text-align: center;">За выбранный период проблем не зафиксировано</td></tr>'
    else:
        for inc in summary['incidents']:
            excluded_note = f" · {esc(inc['excluded'])}" if inc['excluded'] else ''
            html_body += f"""
            <tr>
                <td><b>{esc(inc['client'])}</b></td>
                <td>{esc(inc['server'])}</td>
                <td>{esc(inc['name'])}<div class="inc-meta">{_format_incident_moment(inc['clock'])}{excluded_note}</div></td>
                <td class="inc-meta">{esc(inc['comment']) or '-'}</td>
                <td style="text-align: center;">{format_duration(inc['downtime_sec'])}</td>
            </tr>
            """

    html_body += """
            </tbody>
        </table>

        <p class="footer">Этот же отчет в формате Excel прикреплен во вложении к данному письму.</p>
        <p class="footer">С уважением,<br>Служба мониторинга ИТ-инфраструктуры</p>
    </body>
    </html>
    """

    # Создание письма
    msg = MIMEMultipart()
    msg['From'] = smtp_from
    msg['To'] = recipient
    msg['Subject'] = subject

    msg.attach(MIMEText(html_body, 'html'))

    # Генерация и прикрепление Excel
    excel_stream = generate_excel_report(start_time, end_time)
    attachment = MIMEBase('application', 'octet-stream')
    attachment.set_payload(excel_stream.read())
    encoders.encode_base64(attachment)

    filename = f"SLI_SLA_Report_{date_start_str}_{date_end_str}.xlsx"
    attachment.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(attachment)

    # Отправка
    server = smtplib.SMTP(smtp_host, smtp_port)
    if use_tls:
        server.starttls()
    if smtp_user and smtp_password:
        server.login(smtp_user, smtp_password)
    server.sendmail(smtp_from, recipient, msg.as_string())
    server.quit()

def send_telegram_alert(text):
    """
    Отправляет мгновенное текстовое оповещение в Telegram-канал/чат.
    """
    settings = database.get_settings()
    enabled = settings.get('telegram_enabled') == '1'
    token = settings.get('telegram_bot_token')
    chat_id = settings.get('telegram_chat_id')

    if not enabled or not token or not chat_id:
        return False

    return _send_telegram_message(token, chat_id, text)

def _send_telegram_message(token, chat_id, text):
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "HTML"
    }
    try:
        response = requests.post(url, json=payload, timeout=10)
        return response.status_code == 200
    except Exception:
        return False

def _send_telegram_blocks(token, chat_id, blocks, limit=3800):
    """
    Отправляет список текстовых блоков, склеивая их в сообщения
    не длиннее лимита Telegram (4096 символов) и разбивая на несколько сообщений.
    """
    chunk = ''
    for block in blocks:
        if chunk and len(chunk) + len(block) + 2 > limit:
            _send_telegram_message(token, chat_id, chunk.rstrip())
            chunk = ''
        chunk += block + '\n'
    if chunk.strip():
        _send_telegram_message(token, chat_id, chunk.rstrip())

def send_telegram_report(start_time, end_time):
    """
    Отправляет в Telegram отчет в структуре печатной версии (PDF):
    KPI, топ-20 проблемных серверов, все проблемы по серверам + Excel-файл.
    """
    settings = database.get_settings()
    enabled = settings.get('telegram_enabled') == '1'
    token = settings.get('telegram_bot_token')
    chat_id = settings.get('telegram_chat_id')

    if not enabled or not token or not chat_id:
        return False

    summary = build_report_summary(start_time, end_time)
    esc = html_lib.escape

    date_start_str = datetime.datetime.fromtimestamp(start_time).strftime('%d.%m.%Y')
    date_end_str = datetime.datetime.fromtimestamp(end_time).strftime('%d.%m.%Y')

    blocks = []
    blocks.append(f"📊 <b>Отчет SLI/SLA за {date_start_str} - {date_end_str}</b>\n")
    blocks.append(
        f"Средний SLA (по серверам): <b>{summary['avg_sla']}%</b>\n"
        f"Всего инцидентов: <b>{summary['total_incidents']}</b>\n"
        f"Ср. реакция (MTTD): <code>{format_duration(summary['avg_mttd'])}</code>\n"
        f"Ср. решение (MTTR): <code>{format_duration(summary['avg_mttr'])}</code>\n"
    )

    blocks.append("🏆 <b>Топ-20 проблемных серверов (по SLA)</b>")
    if not summary['top_servers']:
        blocks.append("Нет данных")
    else:
        for idx, srv in enumerate(summary['top_servers'], 1):
            emoji = "🟢" if srv['sla_percent'] >= 99.5 else ("🟡" if srv['sla_percent'] >= 98.0 else "🔴")
            blocks.append(
                f"{emoji} {idx}. <b>{esc(srv['client'])}</b> / {esc(srv['name'])} — <b>{srv['sla_percent']}%</b>\n"
                f"     простой <code>{format_duration(srv['downtime_sec'])}</code>, сбоев {srv['incidents_count']}"
            )

    blocks.append("\n⚠️ <b>Все проблемы по серверам за период</b>")
    if not summary['incidents']:
        blocks.append("За выбранный период проблем не зафиксировано")
    else:
        for inc in summary['incidents']:
            excluded_note = f" ({esc(inc['excluded'])})" if inc['excluded'] else ''
            block = (
                f"▪️ <b>{esc(inc['client'])}</b> / {esc(inc['server'])}\n"
                f"     {esc(inc['name'])} — {_format_incident_moment(inc['clock'])}, "
                f"простой <code>{format_duration(inc['downtime_sec'])}</code>{excluded_note}"
            )
            if inc['comment']:
                block += f"\n     💬 {esc(inc['comment'])}"
            blocks.append(block)

    try:
        _send_telegram_blocks(token, chat_id, blocks)

        # Так же сгенерируем и отправим Excel-отчет
        excel_stream = generate_excel_report(start_time, end_time)
        doc_url = f"https://api.telegram.org/bot{token}/sendDocument"
        files = {
            'document': (f"SLI_SLA_Report_{date_start_str}_{date_end_str}.xlsx", excel_stream, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        data = {
            'chat_id': chat_id,
            'caption': f"Документ: Отчет SLI/SLA ({date_start_str} - {date_end_str})"
        }
        requests.post(doc_url, data=data, files=files, timeout=10)
        return True
    except Exception:
        return False
